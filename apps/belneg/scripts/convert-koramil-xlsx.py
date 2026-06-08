#!/usr/bin/env python3
"""Convert DAFTAR KORAMIL INDONESIA TA2025.xlsx into a clean JSON for ingestion.

Usage:
    python3 apps/belneg/scripts/convert-koramil-xlsx.py
    python3 apps/belneg/scripts/convert-koramil-xlsx.py --xlsx <path> --out <path>

Output: apps/belneg/data/koramil.json — array of { kodam_name, korem_name,
kodim_name, koramil_name, alamat, danramil_name, pangkat, phone, bentuk }.

Skips header/divider rows. Strips obvious whitespace/cell-merge noise.
"""

from __future__ import annotations
import argparse, json, re
from pathlib import Path
from openpyxl import load_workbook

REPO = Path(__file__).resolve().parents[3]

def clean(s):
    if s is None: return None
    s = str(s).strip()
    return s if s else None

def main():
    p = argparse.ArgumentParser(description=__doc__)
    p.add_argument("--xlsx", type=Path, default=REPO / "KODAMKOREMKODIM" / "DAFTAR KORAMIL INDONESIA TA2025.xlsx")
    p.add_argument("--out",  type=Path, default=REPO / "apps" / "belneg" / "data" / "koramil.json")
    args = p.parse_args()

    wb = load_workbook(args.xlsx, read_only=True, data_only=True)
    ws = wb["MASTER KORAMIL"]
    rows = list(ws.iter_rows(values_only=True))

    # Real data rows have integer NO in col 0
    data = []
    for r in rows:
        if not isinstance(r[0], (int, float)): continue
        kodam   = clean(r[1])
        korem   = clean(r[2])
        kodim   = clean(r[3])
        koramil = clean(r[4])
        if not (kodam and korem and kodim and koramil):
            continue
        # Filter out section-divider-like rows where koramil is just "KORAMIL" alone
        if koramil.upper() == "KORAMIL":
            continue
        data.append({
            "no":            int(r[0]),
            "kodam_name":    kodam,
            "korem_name":    korem,
            "kodim_name":    kodim,
            "koramil_name":  koramil,
            "alamat":        clean(r[5]),
            "jabatan":       clean(r[6]),
            "danramil_name": clean(r[7]),
            "pangkat":       clean(r[8]),
            "phone_office":  clean(r[9]),
            "phone_mobile":  clean(r[10]),
            "bentuk":        clean(r[11]),
        })

    args.out.parent.mkdir(parents=True, exist_ok=True)
    args.out.write_text(json.dumps(data, ensure_ascii=False, indent=1))

    # Stats
    print(f"✓ {len(data)} koramil rows written → {args.out}")
    by_kodam = {}
    for r in data: by_kodam[r["kodam_name"]] = by_kodam.get(r["kodam_name"], 0) + 1
    print(f"  Distinct KODAM: {len(by_kodam)}")
    for k, n in sorted(by_kodam.items(), key=lambda kv: -kv[1])[:5]:
        print(f"    {k}: {n}")
    with_alamat = sum(1 for r in data if r["alamat"])
    with_phone  = sum(1 for r in data if r["phone_mobile"] or r["phone_office"])
    print(f"  with alamat: {with_alamat}/{len(data)} ({100*with_alamat/len(data):.0f}%)")
    print(f"  with phone:  {with_phone}/{len(data)} ({100*with_phone/len(data):.0f}%)")

if __name__ == "__main__":
    main()
